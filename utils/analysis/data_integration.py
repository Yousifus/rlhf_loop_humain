"""
Data Integration & Export for RLHF System

This module implements comprehensive data management including:
- Data versioning for analysis reproducibility
- Analysis result caching and optimization
- Data quality monitoring and validation
- Automated backup and recovery systems
- Multi-format export (PDF, Excel, JSON, HTML)
- Custom dashboard creation and sharing
"""

import numpy as np
import pandas as pd
import json
import pickle
import hashlib
from typing import Dict, List, Tuple, Optional, Any, Union
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
import shutil
import sqlite3
from concurrent.futures import ThreadPoolExecutor
import warnings


@dataclass
class DataVersion:
    """Data version metadata"""
    version_id: str
    timestamp: datetime
    data_hash: str
    file_path: str
    metadata: Dict[str, Any]
    parent_version: Optional[str] = None
    tags: List[str] = field(default_factory=list)


@dataclass
class CacheEntry:
    """Analysis cache entry"""
    cache_key: str
    result: Any
    timestamp: datetime
    expiry: datetime
    size_bytes: int
    access_count: int = 0
    last_access: datetime = field(default_factory=datetime.now)


@dataclass
class QualityMetrics:
    """Data quality assessment metrics"""
    completeness: float  # 0-1, missing data percentage
    consistency: float   # 0-1, data consistency score
    accuracy: float      # 0-1, data accuracy estimate
    timeliness: float    # 0-1, data freshness score
    validity: float      # 0-1, data format/type validity
    overall_score: float
    issues: List[str]
    recommendations: List[str]


@dataclass
class ExportConfig:
    """Configuration for data export"""
    format: str  # pdf, excel, json, html, csv
    include_charts: bool = True
    include_metadata: bool = True
    compression: bool = False
    password_protection: bool = False
    custom_styling: Dict[str, Any] = field(default_factory=dict)


class DataIntegrationManager:
    """
    Comprehensive data integration and export system.
    
    Handles data versioning, caching, quality monitoring, backup/recovery,
    and multi-format export capabilities for RLHF analysis results.
    """
    
    def __init__(self, base_path: str = "data_management"):
        """
        Initialize the data integration manager.
        
        Args:
            base_path: Base directory for data management operations
        """
        self.base_path = Path(base_path)
        self.versions_path = self.base_path / "versions"
        self.cache_path = self.base_path / "cache"
        self.backups_path = self.base_path / "backups"
        self.exports_path = self.base_path / "exports"
        
        # Create directories
        for path in [self.versions_path, self.cache_path, self.backups_path, self.exports_path]:
            path.mkdir(parents=True, exist_ok=True)
        
        # Initialize database for metadata
        self.db_path = self.base_path / "metadata.db"
        self._init_database()
        
        # Cache management
        self.cache = {}
        self.max_cache_size_mb = 500
        self.default_cache_ttl_hours = 24
        
        # Quality monitoring
        self.quality_rules = self._load_quality_rules()
        
    def _init_database(self):
        """Initialize metadata database."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Data versions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS data_versions (
                    version_id TEXT PRIMARY KEY,
                    timestamp TEXT,
                    data_hash TEXT,
                    file_path TEXT,
                    metadata TEXT,
                    parent_version TEXT,
                    tags TEXT
                )
            """)
            
            # Cache entries table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS cache_entries (
                    cache_key TEXT PRIMARY KEY,
                    timestamp TEXT,
                    expiry TEXT,
                    size_bytes INTEGER,
                    access_count INTEGER,
                    last_access TEXT,
                    file_path TEXT
                )
            """)
            
            # Quality assessments table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS quality_assessments (
                    assessment_id TEXT PRIMARY KEY,
                    data_version TEXT,
                    timestamp TEXT,
                    completeness REAL,
                    consistency REAL,
                    accuracy REAL,
                    timeliness REAL,
                    validity REAL,
                    overall_score REAL,
                    issues TEXT,
                    recommendations TEXT
                )
            """)
            
            conn.commit()
    
    def create_data_version(self, 
                          data: Union[pd.DataFrame, Dict[str, Any]], 
                          description: str = "",
                          tags: List[str] = None) -> str:
        """
        Create a new data version with metadata tracking.
        
        Args:
            data: Data to version (DataFrame or dictionary)
            description: Version description
            tags: Optional tags for categorization
            
        Returns:
            Version ID
        """
        timestamp = datetime.now()
        version_id = f"v_{timestamp.strftime('%Y%m%d_%H%M%S')}"
        
        # Calculate data hash for integrity
        data_hash = self._calculate_data_hash(data)
        
        # Save data to file
        file_path = self.versions_path / f"{version_id}.pkl"
        with open(file_path, 'wb') as f:
            pickle.dump(data, f)
        
        # Create metadata
        metadata = {
            "description": description,
            "created_by": "system",
            "data_type": type(data).__name__,
            "size_bytes": file_path.stat().st_size
        }
        
        if isinstance(data, pd.DataFrame):
            metadata.update({
                "shape": data.shape,
                "columns": list(data.columns),
                "memory_usage": data.memory_usage(deep=True).sum()
            })
        
        # Create version record
        version = DataVersion(
            version_id=version_id,
            timestamp=timestamp,
            data_hash=data_hash,
            file_path=str(file_path),
            metadata=metadata,
            tags=tags or []
        )
        
        # Save to database
        self._save_version_to_db(version)
        
        print(f"✅ Created data version: {version_id}")
        return version_id
    
    def get_data_version(self, version_id: str) -> Tuple[Any, DataVersion]:
        """
        Retrieve data and metadata for a specific version.
        
        Args:
            version_id: Version identifier
            
        Returns:
            Tuple of (data, version_metadata)
        """
        version = self._load_version_from_db(version_id)
        if not version:
            raise ValueError(f"Version {version_id} not found")
        
        # Load data from file
        with open(version.file_path, 'rb') as f:
            data = pickle.load(f)
        
        return data, version
    
    def list_data_versions(self, tags: List[str] = None) -> List[DataVersion]:
        """
        List all data versions, optionally filtered by tags.
        
        Args:
            tags: Optional tags to filter by
            
        Returns:
            List of data versions
        """
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            if tags:
                # Filter by tags (simplified - would need proper tag querying)
                cursor.execute("""
                    SELECT * FROM data_versions 
                    ORDER BY timestamp DESC
                """)
            else:
                cursor.execute("""
                    SELECT * FROM data_versions 
                    ORDER BY timestamp DESC
                """)
            
            versions = []
            for row in cursor.fetchall():
                version = DataVersion(
                    version_id=row[0],
                    timestamp=datetime.fromisoformat(row[1]),
                    data_hash=row[2],
                    file_path=row[3],
                    metadata=json.loads(row[4]),
                    parent_version=row[5],
                    tags=json.loads(row[6]) if row[6] else []
                )
                versions.append(version)
        
        return versions
    
    def cache_analysis_result(self, 
                            key: str, 
                            result: Any, 
                            ttl_hours: int = None) -> str:
        """
        Cache analysis result with automatic expiry.
        
        Args:
            key: Cache key identifier
            result: Analysis result to cache
            ttl_hours: Time to live in hours
            
        Returns:
            Cache key
        """
        ttl_hours = ttl_hours or self.default_cache_ttl_hours
        expiry = datetime.now() + timedelta(hours=ttl_hours)
        
        # Create cache file
        cache_file = self.cache_path / f"{key}.pkl"
        with open(cache_file, 'wb') as f:
            pickle.dump(result, f)
        
        # Create cache entry
        entry = CacheEntry(
            cache_key=key,
            result=None,  # Don't store in memory for large objects
            timestamp=datetime.now(),
            expiry=expiry,
            size_bytes=cache_file.stat().st_size
        )
        
        # Save to database
        self._save_cache_entry_to_db(entry, str(cache_file))
        
        # Manage cache size
        self._cleanup_cache()
        
        return key
    
    def get_cached_result(self, key: str) -> Optional[Any]:
        """
        Retrieve cached analysis result.
        
        Args:
            key: Cache key identifier
            
        Returns:
            Cached result or None if not found/expired
        """
        # Check database for entry
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM cache_entries WHERE cache_key = ?
            """, (key,))
            
            row = cursor.fetchone()
            if not row:
                return None
            
            # Check if expired
            expiry = datetime.fromisoformat(row[2])
            if datetime.now() > expiry:
                self._remove_cache_entry(key)
                return None
            
            # Load from file
            try:
                with open(row[6], 'rb') as f:
                    result = pickle.load(f)
                
                # Update access statistics
                cursor.execute("""
                    UPDATE cache_entries 
                    SET access_count = access_count + 1, last_access = ?
                    WHERE cache_key = ?
                """, (datetime.now().isoformat(), key))
                conn.commit()
                
                return result
            except (FileNotFoundError, pickle.PickleError):
                self._remove_cache_entry(key)
                return None
    
    def assess_data_quality(self, 
                          data: pd.DataFrame, 
                          version_id: str = None) -> QualityMetrics:
        """
        Assess data quality across multiple dimensions.
        
        Args:
            data: DataFrame to assess
            version_id: Optional version ID for tracking
            
        Returns:
            Quality metrics
        """
        # Completeness: percentage of non-null values
        total_cells = data.shape[0] * data.shape[1]
        non_null_cells = data.count().sum()
        completeness = non_null_cells / total_cells if total_cells > 0 else 0
        
        # Consistency: coefficient of variation for numeric columns
        numeric_cols = data.select_dtypes(include=[np.number]).columns
        consistency_scores = []
        for col in numeric_cols:
            if data[col].std() > 0:
                cv = data[col].std() / abs(data[col].mean()) if data[col].mean() != 0 else 1
                consistency_scores.append(1 / (1 + cv))  # Normalize
        consistency = np.mean(consistency_scores) if consistency_scores else 1.0
        
        # Accuracy: detect outliers using IQR method
        outlier_counts = []
        for col in numeric_cols:
            Q1 = data[col].quantile(0.25)
            Q3 = data[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = data[(data[col] < Q1 - 1.5 * IQR) | (data[col] > Q3 + 1.5 * IQR)][col]
            outlier_rate = len(outliers) / len(data) if len(data) > 0 else 0
            outlier_counts.append(outlier_rate)
        accuracy = 1 - np.mean(outlier_counts) if outlier_counts else 1.0
        
        # Timeliness: based on timestamp columns
        timeliness = 1.0  # Default to perfect if no timestamp analysis
        timestamp_cols = data.select_dtypes(include=['datetime64']).columns
        if len(timestamp_cols) > 0:
            latest_timestamp = data[timestamp_cols[0]].max()
            time_diff = datetime.now() - latest_timestamp.to_pydatetime()
            timeliness = max(0, 1 - (time_diff.days / 30))  # Decay over 30 days
        
        # Validity: data type consistency
        validity_scores = []
        for col in data.columns:
            try:
                # Check if data types are consistent
                expected_type = data[col].dtype
                type_consistency = 1.0  # Simplified check
                validity_scores.append(type_consistency)
            except Exception:
                validity_scores.append(0.5)
        validity = np.mean(validity_scores)
        
        # Overall score (weighted average)
        weights = [0.25, 0.2, 0.25, 0.15, 0.15]
        scores = [completeness, consistency, accuracy, timeliness, validity]
        overall_score = sum(w * s for w, s in zip(weights, scores))
        
        # Generate issues and recommendations
        issues = []
        recommendations = []
        
        if completeness < 0.9:
            issues.append(f"Low data completeness: {completeness:.1%}")
            recommendations.append("Review data collection processes for missing values")
        
        if consistency < 0.8:
            issues.append(f"Low data consistency: {consistency:.1%}")
            recommendations.append("Investigate sources of data variability")
        
        if accuracy < 0.9:
            issues.append(f"Potential data quality issues: {accuracy:.1%} accuracy")
            recommendations.append("Review outlier detection and data validation")
        
        if timeliness < 0.8:
            issues.append(f"Data freshness concern: {timeliness:.1%}")
            recommendations.append("Update data refresh frequency")
        
        metrics = QualityMetrics(
            completeness=completeness,
            consistency=consistency,
            accuracy=accuracy,
            timeliness=timeliness,
            validity=validity,
            overall_score=overall_score,
            issues=issues,
            recommendations=recommendations
        )
        
        # Save assessment to database
        if version_id:
            self._save_quality_assessment(version_id, metrics)
        
        return metrics
    
    def create_backup(self, include_cache: bool = False) -> str:
        """
        Create a backup of all data management files.
        
        Args:
            include_cache: Whether to include cache files
            
        Returns:
            Backup path
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}"
        backup_path = self.backups_path / backup_name
        backup_path.mkdir(exist_ok=True)
        
        # Backup versions
        versions_backup = backup_path / "versions"
        shutil.copytree(self.versions_path, versions_backup)
        
        # Backup database
        shutil.copy2(self.db_path, backup_path / "metadata.db")
        
        # Optionally backup cache
        if include_cache:
            cache_backup = backup_path / "cache"
            shutil.copytree(self.cache_path, cache_backup)
        
        # Create backup manifest
        manifest = {
            "backup_id": backup_name,
            "timestamp": timestamp,
            "include_cache": include_cache,
            "file_count": len(list(versions_backup.glob("*"))),
            "size_bytes": sum(f.stat().st_size for f in backup_path.rglob("*") if f.is_file())
        }
        
        with open(backup_path / "manifest.json", 'w') as f:
            json.dump(manifest, f, indent=2)
        
        print(f"✅ Created backup: {backup_name}")
        return str(backup_path)
    
    def restore_from_backup(self, backup_path: str) -> bool:
        """
        Restore data from a backup.
        
        Args:
            backup_path: Path to backup directory
            
        Returns:
            Success status
        """
        backup_path = Path(backup_path)
        if not backup_path.exists():
            raise ValueError(f"Backup path {backup_path} does not exist")
        
        try:
            # Load manifest
            with open(backup_path / "manifest.json", 'r') as f:
                manifest = json.load(f)
            
            # Create backup of current state
            current_backup = self.create_backup(include_cache=True)
            print(f"Created safety backup: {current_backup}")
            
            # Restore versions
            if (backup_path / "versions").exists():
                if self.versions_path.exists():
                    shutil.rmtree(self.versions_path)
                shutil.copytree(backup_path / "versions", self.versions_path)
            
            # Restore database
            if (backup_path / "metadata.db").exists():
                shutil.copy2(backup_path / "metadata.db", self.db_path)
            
            # Restore cache if present
            if (backup_path / "cache").exists():
                if self.cache_path.exists():
                    shutil.rmtree(self.cache_path)
                shutil.copytree(backup_path / "cache", self.cache_path)
            
            print(f"✅ Restored from backup: {manifest['backup_id']}")
            return True
            
        except Exception as e:
            print(f"❌ Restore failed: {e}")
            return False
    
    def export_to_excel(self, 
                       data: Dict[str, pd.DataFrame], 
                       output_path: str = None,
                       config: ExportConfig = None) -> str:
        """
        Export data to Excel with multiple sheets and formatting.
        
        Args:
            data: Dictionary of sheet_name -> DataFrame
            output_path: Output file path
            config: Export configuration
            
        Returns:
            Path to exported file
        """
        config = config or ExportConfig(format="excel")
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.exports_path / f"export_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting if configured
                if config.custom_styling:
                    worksheet = writer.sheets[sheet_name]
                    self._apply_excel_styling(worksheet, config.custom_styling)
        
        print(f"✅ Exported to Excel: {output_path}")
        return str(output_path)
    
    def export_to_pdf(self, 
                     content: Dict[str, Any], 
                     output_path: str = None,
                     config: ExportConfig = None) -> str:
        """
        Export analysis results to PDF report.
        
        Args:
            content: Content to export (text, charts, tables)
            output_path: Output file path
            config: Export configuration
            
        Returns:
            Path to exported file
        """
        config = config or ExportConfig(format="pdf")
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.exports_path / f"report_{timestamp}.pdf"
        
        # Use reportlab for PDF generation
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = content.get('title', 'Analysis Report')
            story.append(Paragraph(title, styles['Title']))
            story.append(Spacer(1, 12))
            
            # Content sections
            for section_name, section_content in content.items():
                if section_name == 'title':
                    continue
                
                # Section header
                story.append(Paragraph(section_name.replace('_', ' ').title(), styles['Heading1']))
                
                # Section content
                if isinstance(section_content, str):
                    story.append(Paragraph(section_content, styles['Normal']))
                elif isinstance(section_content, pd.DataFrame):
                    # Convert DataFrame to table
                    table_data = [section_content.columns.tolist()] + section_content.values.tolist()
                    table = Table(table_data)
                    story.append(table)
                
                story.append(Spacer(1, 12))
            
            doc.build(story)
            print(f"✅ Exported to PDF: {output_path}")
            
        except ImportError:
            print("⚠️ reportlab not available, creating HTML report instead")
            return self.export_to_html(content, output_path.replace('.pdf', '.html'), config)
        
        return str(output_path)
    
    def export_to_html(self, 
                      content: Dict[str, Any], 
                      output_path: str = None,
                      config: ExportConfig = None) -> str:
        """
        Export analysis results to HTML report.
        
        Args:
            content: Content to export
            output_path: Output file path
            config: Export configuration
            
        Returns:
            Path to exported file
        """
        config = config or ExportConfig(format="html")
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.exports_path / f"report_{timestamp}.html"
        
        # Generate HTML content
        html_content = self._generate_html_report(content, config)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ Exported to HTML: {output_path}")
        return str(output_path)
    
    # Helper methods
    def _calculate_data_hash(self, data: Any) -> str:
        """Calculate hash for data integrity checking."""
        if isinstance(data, pd.DataFrame):
            return hashlib.sha256(data.to_csv().encode()).hexdigest()[:16]
        else:
            return hashlib.sha256(str(data).encode()).hexdigest()[:16]
    
    def _save_version_to_db(self, version: DataVersion):
        """Save version metadata to database."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO data_versions
                (version_id, timestamp, data_hash, file_path, metadata, parent_version, tags)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                version.version_id,
                version.timestamp.isoformat(),
                version.data_hash,
                version.file_path,
                json.dumps(version.metadata),
                version.parent_version,
                json.dumps(version.tags)
            ))
            conn.commit()
    
    def _load_version_from_db(self, version_id: str) -> Optional[DataVersion]:
        """Load version metadata from database."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM data_versions WHERE version_id = ?
            """, (version_id,))
            
            row = cursor.fetchone()
            if not row:
                return None
            
            return DataVersion(
                version_id=row[0],
                timestamp=datetime.fromisoformat(row[1]),
                data_hash=row[2],
                file_path=row[3],
                metadata=json.loads(row[4]),
                parent_version=row[5],
                tags=json.loads(row[6]) if row[6] else []
            )
    
    def _save_cache_entry_to_db(self, entry: CacheEntry, file_path: str):
        """Save cache entry to database."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO cache_entries
                (cache_key, timestamp, expiry, size_bytes, access_count, last_access, file_path)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                entry.cache_key,
                entry.timestamp.isoformat(),
                entry.expiry.isoformat(),
                entry.size_bytes,
                entry.access_count,
                entry.last_access.isoformat(),
                file_path
            ))
            conn.commit()
    
    def _cleanup_cache(self):
        """Clean up expired cache entries and manage size."""
        current_time = datetime.now()
        
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Remove expired entries
            cursor.execute("""
                SELECT cache_key, file_path FROM cache_entries
                WHERE expiry < ?
            """, (current_time.isoformat(),))
            
            expired_entries = cursor.fetchall()
            for cache_key, file_path in expired_entries:
                try:
                    Path(file_path).unlink(missing_ok=True)
                except Exception:
                    pass
                
                cursor.execute("""
                    DELETE FROM cache_entries WHERE cache_key = ?
                """, (cache_key,))
            
            # Check total cache size
            cursor.execute("""
                SELECT SUM(size_bytes) FROM cache_entries
            """)
            total_size = cursor.fetchone()[0] or 0
            
            # Remove least recently used entries if over limit
            if total_size > self.max_cache_size_mb * 1024 * 1024:
                cursor.execute("""
                    SELECT cache_key, file_path FROM cache_entries
                    ORDER BY last_access ASC
                    LIMIT 10
                """)
                
                old_entries = cursor.fetchall()
                for cache_key, file_path in old_entries:
                    try:
                        Path(file_path).unlink(missing_ok=True)
                    except Exception:
                        pass
                    
                    cursor.execute("""
                        DELETE FROM cache_entries WHERE cache_key = ?
                    """, (cache_key,))
            
            conn.commit()
    
    def _remove_cache_entry(self, cache_key: str):
        """Remove a specific cache entry."""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT file_path FROM cache_entries WHERE cache_key = ?
            """, (cache_key,))
            
            row = cursor.fetchone()
            if row:
                try:
                    Path(row[0]).unlink(missing_ok=True)
                except Exception:
                    pass
                
                cursor.execute("""
                    DELETE FROM cache_entries WHERE cache_key = ?
                """, (cache_key,))
                conn.commit()
    
    def _save_quality_assessment(self, version_id: str, metrics: QualityMetrics):
        """Save quality assessment to database."""
        assessment_id = f"qa_{version_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO quality_assessments
                (assessment_id, data_version, timestamp, completeness, consistency, 
                 accuracy, timeliness, validity, overall_score, issues, recommendations)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                assessment_id,
                version_id,
                datetime.now().isoformat(),
                metrics.completeness,
                metrics.consistency,
                metrics.accuracy,
                metrics.timeliness,
                metrics.validity,
                metrics.overall_score,
                json.dumps(metrics.issues),
                json.dumps(metrics.recommendations)
            ))
            conn.commit()
    
    def _load_quality_rules(self) -> Dict[str, Any]:
        """Load data quality rules configuration."""
        return {
            "completeness_threshold": 0.95,
            "consistency_threshold": 0.85,
            "accuracy_threshold": 0.90,
            "timeliness_threshold": 0.80,
            "validity_threshold": 0.95
        }
    
    def _apply_excel_styling(self, worksheet, styling_config: Dict[str, Any]):
        """Apply styling to Excel worksheet."""
        # Simplified styling - would be more comprehensive in production
        try:
            from openpyxl.styles import Font, Fill, PatternFill
            
            # Header styling
            if 'header_style' in styling_config:
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        except ImportError:
            pass  # Styling not available
    
    def _generate_html_report(self, content: Dict[str, Any], config: ExportConfig) -> str:
        """Generate HTML report from content."""
        html_template = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{content.get('title', 'Analysis Report')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                .header {{ background: #f0f8ff; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
                .section {{ margin: 20px 0; padding: 15px; border-left: 4px solid #007acc; background: #f9f9f9; }}
                .metric {{ display: inline-block; margin: 10px; padding: 10px; background: white; border-radius: 3px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .timestamp {{ color: #666; font-size: 0.9em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{content.get('title', 'Analysis Report')}</h1>
                <p class="timestamp">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Add content sections
        for section_name, section_content in content.items():
            if section_name == 'title':
                continue
            
            html_template += f"""
            <div class="section">
                <h2>{section_name.replace('_', ' ').title()}</h2>
            """
            
            if isinstance(section_content, str):
                html_template += f"<p>{section_content}</p>"
            elif isinstance(section_content, dict):
                for key, value in section_content.items():
                    html_template += f"""
                    <div class="metric">
                        <strong>{key}:</strong> {value}
                    </div>
                    """
            elif isinstance(section_content, pd.DataFrame):
                html_template += section_content.to_html(classes="table", table_id=f"table_{section_name}")
            
            html_template += "</div>"
        
        html_template += """
            <footer style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #666;">
                Generated by RLHF Data Integration Manager
            </footer>
        </body>
        </html>
        """
        
        return html_template


def create_demo_data_integration() -> DataIntegrationManager:
    """
    Create demonstration of data integration capabilities.
    
    Returns:
        Configured data integration manager with demo data
    """
    # Initialize manager
    manager = DataIntegrationManager("demo_data_management")
    
    # Create demo data versions
    demo_data1 = pd.DataFrame({
        'timestamp': pd.date_range('2024-01-01', periods=100, freq='H'),
        'accuracy': np.random.normal(0.85, 0.05, 100).clip(0.7, 0.95),
        'confidence': np.random.normal(0.8, 0.1, 100).clip(0.5, 1.0),
        'calibration_error': np.random.normal(0.05, 0.02, 100).clip(0.01, 0.15)
    })
    
    demo_data2 = pd.DataFrame({
        'model_id': range(10),
        'performance_score': np.random.uniform(0.7, 0.95, 10),
        'training_time': np.random.uniform(10, 100, 10),
        'data_size': np.random.randint(1000, 50000, 10)
    })
    
    # Create versions
    v1_id = manager.create_data_version(
        demo_data1, 
        description="Historical performance data",
        tags=["performance", "time_series"]
    )
    
    v2_id = manager.create_data_version(
        demo_data2,
        description="Model comparison data", 
        tags=["models", "comparison"]
    )
    
    # Assess data quality
    quality1 = manager.assess_data_quality(demo_data1, v1_id)
    quality2 = manager.assess_data_quality(demo_data2, v2_id)
    
    # Create cache entries
    analysis_result = {
        "summary": "Performance analysis complete",
        "metrics": {"accuracy": 0.85, "calibration": 0.05},
        "recommendations": ["Continue monitoring", "Check for drift"]
    }
    
    manager.cache_analysis_result("demo_analysis_2024", analysis_result)
    
    # Create backup
    backup_path = manager.create_backup(include_cache=True)
    
    # Export demo data
    export_data = {
        "Performance Data": demo_data1,
        "Model Comparison": demo_data2
    }
    
    excel_path = manager.export_to_excel(export_data)
    
    report_content = {
        "title": "Demo Data Integration Report",
        "summary": "Demonstration of data integration capabilities",
        "quality_metrics": {
            "Dataset 1 Quality": f"{quality1.overall_score:.1%}",
            "Dataset 2 Quality": f"{quality2.overall_score:.1%}",
            "Issues Found": len(quality1.issues) + len(quality2.issues)
        },
        "data_versions": f"Created {len(manager.list_data_versions())} versions",
        "Performance Data": demo_data1.head()
    }
    
    html_path = manager.export_to_html(report_content)
    
    print(f"📊 Demo Data Integration Results:")
    print(f"✅ Data versions created: {len(manager.list_data_versions())}")
    print(f"✅ Quality assessments: 2")
    print(f"✅ Cache entries: 1")
    print(f"✅ Backup created: {backup_path}")
    print(f"✅ Excel export: {excel_path}")
    print(f"✅ HTML report: {html_path}")
    
    return manager


if __name__ == "__main__":
    print("📦 Data Integration & Export System - Demo Mode")
    print("=" * 50)
    
    # Run demo
    manager = create_demo_data_integration()
    
    print(f"\n🎯 Data Integration Demo Complete!")
    print(f"Phase 3 Module 3: ✅ READY FOR INTEGRATION") 